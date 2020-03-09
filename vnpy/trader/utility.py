"""
General utility functions.
"""

import json
import logging
import sys
import os
import csv
import re
from pathlib import Path
from typing import Callable, Dict
from decimal import Decimal
from math import floor, ceil
from time import time
from datetime import datetime, timedelta
from functools import wraps, lru_cache
import numpy as np
import talib

from .object import BarData, TickData
from .constant import Exchange, Interval

log_formatter = logging.Formatter('[%(asctime)s] %(message)s')


def func_time(over_ms: int = 0):
    """
    简单记录执行时间
    :param :over_ms 超过多少毫秒, 提示信息
    :return:
    """

    def run(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            start = time()
            result = func(*args, **kwargs)
            end = time()
            execute_ms = (int(round(end * 1000))) - (int(round(start * 1000)))
            if execute_ms > over_ms:
                print('{} took {} ms'.format(func.__qualname__, execute_ms))
            return result

        return wrapper

    return run


@lru_cache()
def get_underlying_symbol(symbol: str):
    """
    取得合约的短号.  rb2005 => rb
    :param symbol:
    :return: 短号
    """
    # 套利合约
    if symbol.find(' ') != -1:
        # 排除SP SPC SPD
        s = symbol.split(' ')
        if len(s) < 2:
            return symbol
        symbol = s[1]

        # 只提取leg1合约
        if symbol.find('&') != -1:
            s = symbol.split('&')
            if len(s) < 2:
                return symbol
            symbol = s[0]

    p = re.compile(r"([A-Z]+)[0-9]+", re.I)
    underlying_symbol = p.match(symbol)

    if underlying_symbol is None:
        return symbol

    return underlying_symbol.group(1)


@lru_cache()
def get_stock_exchange(code, vn=True):
    """根据股票代码，获取交易所"""
    # vn：取EXCHANGE_SSE 和 EXCHANGE_SZSE
    code = str(code)
    if len(code) < 6:
        return ''

    market_id = 0  # 缺省深圳
    code = str(code)
    if code[0] in ['5', '6', '9'] or code[:3] in ["009", "126", "110", "201", "202", "203", "204"]:
        market_id = 1  # 上海
    try:
        from vnpy.trader.constant import Exchange
        if vn:
            return Exchange.SSE.value if market_id == 1 else Exchange.SZSE.value
        else:
            return 'XSHG' if market_id == 1 else 'XSHE'
    except Exception as ex:
        print(u'加载数据异常:{}'.format(str(ex)))

    return ''


@lru_cache()
def get_full_symbol(symbol: str):
    """
    获取全路径得合约名称, MA005 => MA2005, j2005 => j2005
    """
    if symbol.endswith('SPD'):
        return symbol

    underlying_symbol = get_underlying_symbol(symbol)
    if underlying_symbol == symbol:
        return symbol

    symbol_month = symbol.replace(underlying_symbol, '')
    if len(symbol_month) == 3:
        if symbol_month[0] == '0':
            # 支持2020年合约
            return '{0}2{1}'.format(underlying_symbol, symbol_month)
        else:
            return '{0}1{1}'.format(underlying_symbol, symbol_month)
    else:
        return symbol


def get_real_symbol_by_exchange(full_symbol, vn_exchange):
    """根据交易所，返回真实合约"""
    if vn_exchange == Exchange.CFFEX:
        return full_symbol.upper()

    if vn_exchange in [Exchange.DCE, Exchange.SHFE, Exchange.INE]:
        return full_symbol.lower()

    if vn_exchange == Exchange.CZCE:
        underlying_symbol = get_underlying_symbol(full_symbol).upper()
        yearmonth_len = len(full_symbol) - len(underlying_symbol) - 1
        return underlying_symbol.upper() + full_symbol[-yearmonth_len:]

    return full_symbol


def get_trading_date(dt: datetime = None):
    """
    根据输入的时间，返回交易日的日期
    :param dt:
    :return:
    """
    if dt is None:
        dt = datetime.now()

    if dt.isoweekday() in [6, 7]:
        # 星期六,星期天=>星期一
        return (dt + timedelta(days=8 - dt.isoweekday())).strftime('%Y-%m-%d')

    if dt.hour >= 20:
        if dt.isoweekday() == 5:
            # 星期五=》星期一
            return (dt + timedelta(days=3)).strftime('%Y-%m-%d')
        else:
            # 第二天
            return (dt + timedelta(days=1)).strftime('%Y-%m-%d')
    else:
        return dt.strftime('%Y-%m-%d')


def extract_vt_symbol(vt_symbol: str):
    """
    :return: (symbol, exchange)
    """
    symbol, exchange_str = vt_symbol.split(".")
    return symbol, Exchange(exchange_str)


def generate_vt_symbol(symbol: str, exchange: Exchange):
    """
    return vt_symbol
    """
    return f"{symbol}.{exchange.value}"


def format_number(n):
    """格式化数字到字符串"""
    rn = round(n, 2)  # 保留两位小数
    return format(rn, ',')  # 加上千分符


def _get_trader_dir(temp_name: str):
    """
    Get path where trader is running in.
    """
    # by incenselee
    # 原方法，当前目录必须自建.vntrader子目录，否则在用户得目录下创建
    # 为兼容多账号管理，取消此方法。
    return Path.cwd(), Path.cwd()

    cwd = Path.cwd()
    temp_path = cwd.joinpath(temp_name)

    # If .vntrader folder exists in current working directory,
    # then use it as trader running path.
    if temp_path.exists():
        return cwd, temp_path

    # Otherwise use home path of system.
    home_path = Path.home()
    temp_path = home_path.joinpath(temp_name)

    # Create .vntrader folder under home path if not exist.
    if not temp_path.exists():
        temp_path.mkdir()

    return home_path, temp_path


TRADER_DIR, TEMP_DIR = _get_trader_dir(".vntrader")
if TRADER_DIR not in sys.path:
    sys.path.append(str(TRADER_DIR))
    print(f'sys.path append: {str(TRADER_DIR)}')


def get_file_path(filename: str):
    """
    Get path for temp file with filename.
    """
    return TEMP_DIR.joinpath(filename)


def get_folder_path(folder_name: str):
    """
    Get path for temp folder with folder name.
    """
    folder_path = TEMP_DIR.joinpath(folder_name)
    if not folder_path.exists():
        folder_path.mkdir()
    return folder_path


def get_icon_path(filepath: str, ico_name: str):
    """
    Get path for icon file with ico name.
    """
    ui_path = Path(filepath).parent
    icon_path = ui_path.joinpath("ico", ico_name)
    return str(icon_path)


def load_json(filename: str, auto_save: bool = True):
    """
    Load data from json file in temp path.
    """
    filepath = get_file_path(filename)

    if filepath.exists():
        with open(filepath, mode="r", encoding="UTF-8") as f:
            data = json.load(f)
        return data
    else:
        if auto_save:
            save_json(filename, {})
        return {}


def save_json(filename: str, data: dict):
    """
    Save data into json file in temp path.
    """
    filepath = get_file_path(filename)
    with open(filepath, mode="w+", encoding="UTF-8") as f:
        json.dump(
            data,
            f,
            indent=4,
            ensure_ascii=False
        )


def round_to(value: float, target: float) -> float:
    """
    Round price to price tick value.
    """
    value = Decimal(str(value))
    target = Decimal(str(target))
    rounded = float(int(round(value / target)) * target)
    return rounded


def floor_to(value: float, target: float) -> float:
    """
    Similar to math.floor function, but to target float number.
    """
    value = Decimal(str(value))
    target = Decimal(str(target))
    result = float(int(floor(value / target)) * target)
    return result


def ceil_to(value: float, target: float) -> float:
    """
    Similar to math.ceil function, but to target float number.
    """
    value = Decimal(str(value))
    target = Decimal(str(target))
    result = float(int(ceil(value / target)) * target)
    return result


def print_dict(d: dict):
    """返回dict的字符串类型"""
    return '\n'.join([f'{key}:{d[key]}' for key in sorted(d.keys())])


def append_data(file_name: str, dict_data: dict, field_names: list = []):
    """
    添加数据到csv文件中
    :param file_name:  csv的文件全路径
    :param dict_data:  OrderedDict
    :return:
    """
    dict_fieldnames = sorted(list(dict_data.keys())) if len(field_names) == 0 else field_names

    try:
        if not os.path.exists(file_name):
            print(u'create csv file:{}'.format(file_name))
            with open(file_name, 'a', encoding='utf8', newline='\n') as csvWriteFile:
                writer = csv.DictWriter(f=csvWriteFile, fieldnames=dict_fieldnames, dialect='excel')
                print(u'write csv header:{}'.format(dict_fieldnames))
                writer.writeheader()
                writer.writerow(dict_data)
        else:
            with open(file_name, 'a', encoding='utf8', newline='\n') as csvWriteFile:
                writer = csv.DictWriter(f=csvWriteFile, fieldnames=dict_fieldnames, dialect='excel',
                                        extrasaction='ignore')
                writer.writerow(dict_data)
    except Exception as ex:
        print(u'append_data exception:{}'.format(str(ex)), file=sys.stderr)


def import_module_by_str(import_module_name):
    """
    动态导入模块/函数
    :param import_module_name:
    :return:
    """
    import traceback
    from importlib import import_module, reload

    # 参数检查
    if len(import_module_name) == 0:
        print('import_module_by_str parameter error,return None')
        return None

    print('trying to import {}'.format(import_module_name))
    try:
        import_name = str(import_module_name).replace(':', '.')
        modules = import_name.split('.')
        if len(modules) == 1:
            mod = import_module(modules[0])
            return mod
        else:
            loaded_modules = '.'.join(modules[0:-1])
            print('import {}'.format(loaded_modules))
            mod = import_module(loaded_modules)

            comp = modules[-1]
            if not hasattr(mod, comp):
                loaded_modules = '.'.join([loaded_modules, comp])
                print('realod {}'.format(loaded_modules))
                mod = reload(loaded_modules)
            else:
                print('from {} import {}'.format(loaded_modules, comp))
                mod = getattr(mod, comp)
            return mod

    except Exception as ex:
        print('import {} fail,{},{}'.format(import_module_name, str(ex), traceback.format_exc()))

        return None


def save_df_to_excel(file_name, sheet_name, df):
    """
    保存dataframe到execl
    :param file_name: 保存的excel文件名
    :param sheet_name: 保存的sheet
    :param df: dataframe
    :return: True/False
    """
    if file_name is None or sheet_name is None or df is None:
        return False

    # ----------------------------- 扩展的功能 ---------
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        # from openpyxl.drawing.image import Image
    except:  # noqa
        print(u'can not import openpyxl', file=sys.stderr)

    if 'openpyxl' not in sys.modules:
        print(u'can not import openpyxl', file=sys.stderr)
        return False

    try:
        ws = None

        try:
            # 读取文件
            wb = openpyxl.load_workbook(file_name)
        except:  # noqa
            # 创建一个excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
        try:
            # 定位WorkSheet
            if ws is None:
                ws = wb[sheet_name]
        except:  # noqa
            # 创建一个WorkSheet
            ws = wb.create_sheet()
            ws.title = sheet_name

        rows = dataframe_to_rows(df)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Save the workbook
        wb.save(file_name)
        wb.close()
    except Exception as ex:
        import traceback
        print(u'save_df_to_excel exception:{}'.format(str(ex)), traceback.format_exc(), file=sys.stderr)


def save_text_to_excel(file_name, sheet_name, text):
    """
    保存文本文件到excel
    :param file_name:
    :param sheet_name:
    :param text:
    :return:
    """
    if file_name is None or len(sheet_name) == 0 or len(text) == 0:
        return False

    # ----------------------------- 扩展的功能 ---------
    try:
        import openpyxl
        # from openpyxl.utils.dataframe import dataframe_to_rows
        # from openpyxl.drawing.image import Image
    except:  # noqa
        print(u'can not import openpyxl', file=sys.stderr)

    if 'openpyxl' not in sys.modules:
        return False

    try:
        ws = None
        try:
            # 读取文件
            wb = openpyxl.load_workbook(file_name)
        except:  # noqa
            # 创建一个excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
        try:
            # 定位WorkSheet
            if ws is None:
                ws = wb[sheet_name]
        except:  # noqa
            # 创建一个WorkSheet
            ws = wb.create_sheet()
            ws.title = sheet_name

        # 设置宽度，自动换行方式
        ws.column_dimensions["A"].width = 120
        ws['A2'].alignment = openpyxl.styles.Alignment(wrapText=True)
        ws['A2'].value = text

        # Save the workbook
        wb.save(file_name)
        wb.close()
        return True
    except Exception as ex:
        import traceback
        print(u'save_text_to_excel exception:{}'.format(str(ex)), traceback.format_exc(), file=sys.stderr)
        return False


def save_images_to_excel(file_name, sheet_name, image_names):
    """
    # 保存图形文件到excel
    :param file_name: excel文件名
    :param sheet_name: workSheet
    :param image_names: 图像文件名列表
    :return:
    """
    if file_name is None or len(sheet_name) == 0 or len(image_names) == 0:
        return False
    # ----------------------------- 扩展的功能 ---------
    try:
        import openpyxl
        # from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.drawing.image import Image
    except Exception as ex:
        print(f'can not import openpyxl:{str(ex)}', file=sys.stderr)

    if 'openpyxl' not in sys.modules:
        return False
    try:
        ws = None

        try:
            # 读取文件
            wb = openpyxl.load_workbook(file_name)
        except:  # noqa
            # 创建一个excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
        try:
            # 定位WorkSheet
            if ws is None:
                ws = wb[sheet_name]
        except Exception as ex:  # noqa
            # 创建一个WorkSheet
            ws = wb.create_sheet()
            ws.title = sheet_name

        i = 1

        for image_name in image_names:
            try:
                # 加载图形文件
                img1 = Image(image_name)

                cell_id = 'A{0}'.format(i)
                ws[cell_id].value = image_name
                cell_id = 'A{0}'.format(i + 1)

                i += 30

                # 添加至对应的WorkSheet中
                ws.add_image(img1, cell_id)
            except Exception as ex:
                print('exception loading image {}, {}'.format(image_name, str(ex)), file=sys.stderr)
                return False

        # Save the workbook
        wb.save(file_name)
        wb.close()
        return True
    except Exception as ex:
        import traceback
        print(u'save_images_to_excel exception:{}'.format(str(ex)), traceback.format_exc(), file=sys.stderr)
        return False


def display_dual_axis(df, columns1, columns2=[], invert_yaxis1=False, invert_yaxis2=False, file_name=None,
                      sheet_name=None,
                      image_name=None):
    """
    显示(保存)双Y轴的走势图
    :param df: DataFrame
    :param columns1:  y1轴
    :param columns2: Y2轴
    :param invert_yaxis1: Y1 轴反转
    :param invert_yaxis2: Y2 轴翻转
    :param file_name:   保存的excel 文件名称
    :param sheet_name:  excel 的sheet
    :param image_name:  保存的image 文件名
    :return:
    """

    import matplotlib
    import matplotlib.pyplot as plt
    matplotlib.rcParams['figure.figsize'] = (20.0, 10.0)

    df1 = df[columns1]
    df1.index = list(range(len(df)))
    fig, ax1 = plt.subplots()
    if invert_yaxis1:
        ax1.invert_yaxis()
    ax1.plot(df1)

    if len(columns2) > 0:
        df2 = df[columns2]
        df2.index = list(range(len(df)))
        ax2 = ax1.twinx()
        if invert_yaxis2:
            ax2.invert_yaxis()
        ax2.plot(df2)

    # 修改x轴得label为时间
    xt = ax1.get_xticks()
    xt2 = [df.index[int(i)] for i in xt[1:-2]]
    xt2.insert(0, '')
    xt2.append('')
    ax1.set_xticklabels(xt2)

    # 是否保存图片到文件
    if image_name is not None:
        fig = plt.gcf()
        fig.savefig(image_name, bbox_inches='tight')

        # 插入图片到指定的excel文件sheet中并保存excel
        if file_name is not None and sheet_name is not None:
            save_images_to_excel(file_name, sheet_name, [image_name])
    else:
        plt.show()


class BarGenerator:
    """
    For:
    1. generating 1 minute bar data from tick data
    2. generateing x minute bar/x hour bar data from 1 minute data

    Notice:
    1. for x minute bar, x must be able to divide 60: 2, 3, 5, 6, 10, 15, 20, 30
    2. for x hour bar, x can be any number
    """

    def __init__(
            self,
            on_bar: Callable,
            window: int = 0,
            on_window_bar: Callable = None,
            interval: Interval = Interval.MINUTE
    ):
        """Constructor"""
        self.bar = None
        self.on_bar = on_bar

        self.interval = interval
        self.interval_count = 0

        self.window = window
        self.window_bar = None
        self.on_window_bar = on_window_bar

        self.last_tick = None
        self.last_bar = None

    def update_tick(self, tick: TickData):
        """
        Update new tick data into generator.
        """
        new_minute = False

        # Filter tick data with 0 last price
        if not tick.last_price:
            return

        if not self.bar:
            new_minute = True
        elif self.bar.datetime.minute != tick.datetime.minute:
            self.bar.datetime = self.bar.datetime.replace(
                second=0, microsecond=0
            )
            self.on_bar(self.bar)

            new_minute = True

        if new_minute:
            self.bar = BarData(
                symbol=tick.symbol,
                exchange=tick.exchange,
                interval=Interval.MINUTE,
                datetime=tick.datetime,
                gateway_name=tick.gateway_name,
                open_price=tick.last_price,
                high_price=tick.last_price,
                low_price=tick.last_price,
                close_price=tick.last_price,
                open_interest=tick.open_interest
            )
        else:
            self.bar.high_price = max(self.bar.high_price, tick.last_price)
            self.bar.low_price = min(self.bar.low_price, tick.last_price)
            self.bar.close_price = tick.last_price
            self.bar.open_interest = tick.open_interest
            self.bar.datetime = tick.datetime

        if self.last_tick:
            volume_change = tick.volume - self.last_tick.volume
            self.bar.volume += max(volume_change, 0)

        self.last_tick = tick

    def update_bar(self, bar: BarData):
        """
        Update 1 minute bar into generator
        """
        # If not inited, creaate window bar object
        if not self.window_bar:
            # Generate timestamp for bar data
            if self.interval == Interval.MINUTE:
                dt = bar.datetime.replace(second=0, microsecond=0)
            else:
                dt = bar.datetime.replace(minute=0, second=0, microsecond=0)

            self.window_bar = BarData(
                symbol=bar.symbol,
                exchange=bar.exchange,
                datetime=dt,
                gateway_name=bar.gateway_name,
                open_price=bar.open_price,
                high_price=bar.high_price,
                low_price=bar.low_price
            )
        # Otherwise, update high/low price into window bar
        else:
            self.window_bar.high_price = max(
                self.window_bar.high_price, bar.high_price)
            self.window_bar.low_price = min(
                self.window_bar.low_price, bar.low_price)

        # Update close price/volume into window bar
        self.window_bar.close_price = bar.close_price
        self.window_bar.volume += int(bar.volume)
        self.window_bar.open_interest = bar.open_interest

        # Check if window bar completed
        finished = False

        if self.interval == Interval.MINUTE:
            # x-minute bar
            if not (bar.datetime.minute + 1) % self.window:
                finished = True
        elif self.interval == Interval.HOUR:
            if self.last_bar and bar.datetime.hour != self.last_bar.datetime.hour:
                # 1-hour bar
                if self.window == 1:
                    finished = True
                # x-hour bar
                else:
                    self.interval_count += 1

                    if not self.interval_count % self.window:
                        finished = True
                        self.interval_count = 0

        if finished:
            self.on_window_bar(self.window_bar)
            self.window_bar = None

        # Cache last bar object
        self.last_bar = bar

    def generate(self):
        """
        Generate the bar data and call callback immediately.
        """
        self.bar.datetime = self.bar.datetime.replace(
            second=0, microsecond=0
        )
        self.on_bar(self.bar)
        self.bar = None


class ArrayManager(object):
    """
    For:
    1. time series container of bar data
    2. calculating technical indicator value
    """

    def __init__(self, size=100):
        """Constructor"""
        self.count = 0
        self.size = size
        self.inited = False

        self.open_array = np.zeros(size)
        self.high_array = np.zeros(size)
        self.low_array = np.zeros(size)
        self.close_array = np.zeros(size)
        self.volume_array = np.zeros(size)

    def update_bar(self, bar):
        """
        Update new bar data into array manager.
        """
        self.count += 1
        if not self.inited and self.count >= self.size:
            self.inited = True

        self.open_array[:-1] = self.open_array[1:]
        self.high_array[:-1] = self.high_array[1:]
        self.low_array[:-1] = self.low_array[1:]
        self.close_array[:-1] = self.close_array[1:]
        self.volume_array[:-1] = self.volume_array[1:]

        self.open_array[-1] = bar.open_price
        self.high_array[-1] = bar.high_price
        self.low_array[-1] = bar.low_price
        self.close_array[-1] = bar.close_price
        self.volume_array[-1] = bar.volume

    @property
    def open(self):
        """
        Get open price time series.
        """
        return self.open_array

    @property
    def high(self):
        """
        Get high price time series.
        """
        return self.high_array

    @property
    def low(self):
        """
        Get low price time series.
        """
        return self.low_array

    @property
    def close(self):
        """
        Get close price time series.
        """
        return self.close_array

    @property
    def volume(self):
        """
        Get trading volume time series.
        """
        return self.volume_array

    def sma(self, n, array=False):
        """
        Simple moving average.
        """
        result = talib.SMA(self.close, n)
        if array:
            return result
        return result[-1]

    def kama(self, n, array=False):
        """
        KAMA.
        """
        result = talib.KAMA(self.close, n)
        if array:
            return result
        return result[-1]

    def wma(self, n, array=False):
        """
        WMA.
        """
        result = talib.WMA(self.close, n)
        if array:
            return result
        return result[-1]

    def apo(self, n, array=False):
        """
        APO.
        """
        result = talib.APO(self.close, n)
        if array:
            return result
        return result[-1]

    def cmo(self, n, array=False):
        """
        CMO.
        """
        result = talib.CMO(self.close, n)
        if array:
            return result
        return result[-1]

    def mom(self, n, array=False):
        """
        MOM.
        """
        result = talib.MOM(self.close, n)
        if array:
            return result
        return result[-1]

    def ppo(self, n, array=False):
        """
        PPO.
        """
        result = talib.PPO(self.close, n)
        if array:
            return result
        return result[-1]

    def roc(self, n, array=False):
        """
        ROC.
        """
        result = talib.ROC(self.close, n)
        if array:
            return result
        return result[-1]

    def rocr(self, n, array=False):
        """
        ROCR.
        """
        result = talib.ROCR(self.close, n)
        if array:
            return result
        return result[-1]

    def rocp(self, n, array=False):
        """
        ROCP.
        """
        result = talib.ROCP(self.close, n)
        if array:
            return result
        return result[-1]

    def rocr_100(self, n, array=False):
        """
        ROCR100.
        """
        result = talib.ROCR100(self.close, n)
        if array:
            return result
        return result[-1]

    def trix(self, n, array=False):
        """
        TRIX.
        """
        result = talib.TRIX(self.close, n)
        if array:
            return result
        return result[-1]

    def std(self, n, array=False):
        """
        Standard deviation
        """
        result = talib.STDDEV(self.close, n)
        if array:
            return result
        return result[-1]

    def obv(self, n, array=False):
        """
        OBV.
        """
        result = talib.OBV(self.close, self.volume)
        if array:
            return result
        return result[-1]

    def cci(self, n, array=False):
        """
        Commodity Channel Index (CCI).
        """
        result = talib.CCI(self.high, self.low, self.close, n)
        if array:
            return result
        return result[-1]

    def atr(self, n, array=False):
        """
        Average True Range (ATR).
        """
        result = talib.ATR(self.high, self.low, self.close, n)
        if array:
            return result
        return result[-1]

    def natr(self, n, array=False):
        """
        NATR.
        """
        result = talib.NATR(self.high, self.low, self.close, n)
        if array:
            return result
        return result[-1]

    def rsi(self, n, array=False):
        """
        Relative Strenght Index (RSI).
        """
        result = talib.RSI(self.close, n)
        if array:
            return result
        return result[-1]

    def macd(self, fast_period, slow_period, signal_period, array=False):
        """
        MACD.
        """
        macd, signal, hist = talib.MACD(
            self.close, fast_period, slow_period, signal_period
        )
        if array:
            return macd, signal, hist
        return macd[-1], signal[-1], hist[-1]

    def adx(self, n, array=False):
        """
        ADX.
        """
        result = talib.ADX(self.high, self.low, self.close, n)
        if array:
            return result
        return result[-1]

    def adxr(self, n, array=False):
        """
        ADXR.
        """
        result = talib.ADXR(self.high, self.low, self.close, n)
        if array:
            return result
        return result[-1]

    def dx(self, n, array=False):
        """
        DX.
        """
        result = talib.DX(self.high, self.low, self.close, n)
        if array:
            return result
        return result[-1]

    def minus_di(self, n, array=False):
        """
        MINUS_DI.
        """
        result = talib.MINUS_DI(self.high, self.low, self.close, n)
        if array:
            return result
        return result[-1]

    def plus_di(self, n, array=False):
        """
        PLUS_DI.
        """
        result = talib.PLUS_DI(self.high, self.low, self.close, n)
        if array:
            return result
        return result[-1]

    def willr(self, n, array=False):
        """
        WILLR.
        """
        result = talib.WILLR(self.high, self.low, self.close, n)
        if array:
            return result
        return result[-1]

    def ultosc(self, array=False):
        """
        Ultimate Oscillator.
        """
        result = talib.ULTOSC(self.high, self.low, self.close)
        if array:
            return result
        return result[-1]

    def trange(self, array=False):
        """
        TRANGE.
        """
        result = talib.TRANGE(self.high, self.low, self.close)
        if array:
            return result
        return result[-1]

    def boll(self, n, dev, array=False):
        """
        Bollinger Channel.
        """
        mid = self.sma(n, array)
        std = self.std(n, array)

        up = mid + std * dev
        down = mid - std * dev

        return up, down

    def keltner(self, n, dev, array=False):
        """
        Keltner Channel.
        """
        mid = self.sma(n, array)
        atr = self.atr(n, array)

        up = mid + atr * dev
        down = mid - atr * dev

        return up, down

    def donchian(self, n, array=False):
        """
        Donchian Channel.
        """
        up = talib.MAX(self.high, n)
        down = talib.MIN(self.low, n)

        if array:
            return up, down
        return up[-1], down[-1]

    def aroon(self, n, array=False):
        """
        Aroon indicator.
        """
        aroon_up, aroon_down = talib.AROON(self.high, self.low, n)

        if array:
            return aroon_up, aroon_down
        return aroon_up[-1], aroon_down[-1]

    def aroonosc(self, n, array=False):
        """
        Aroon Oscillator.
        """
        result = talib.AROONOSC(self.high, self.low, n)

        if array:
            return result
        return result[-1]

    def minus_dm(self, n, array=False):
        """
        MINUS_DM.
        """
        result = talib.MINUS_DM(self.high, self.low, n)

        if array:
            return result
        return result[-1]

    def plus_dm(self, n, array=False):
        """
        PLUS_DM.
        """
        result = talib.PLUS_DM(self.high, self.low, n)

        if array:
            return result
        return result[-1]

    def mfi(self, n, array=False):
        """
        Money Flow Index.
        """
        result = talib.MFI(self.high, self.low, self.close, self.volume, n)
        if array:
            return result
        return result[-1]

    def ad(self, n, array=False):
        """
        AD.
        """
        result = talib.AD(self.high, self.low, self.close, self.volume, n)
        if array:
            return result
        return result[-1]

    def adosc(self, n, array=False):
        """
        ADOSC.
        """
        result = talib.ADOSC(self.high, self.low, self.close, self.volume, n)
        if array:
            return result
        return result[-1]

    def bop(self, array=False):
        result = talib.BOP(self.open, self.high, self.low, self.close)

        if array:
            return result
        return result[-1]


def virtual(func: "callable"):
    """
    mark a function as "virtual", which means that this function can be override.
    any base class should use this or @abstractmethod to decorate all functions
    that can be (re)implemented by subclasses.
    """
    return func


file_handlers: Dict[str, logging.FileHandler] = {}


def _get_file_logger_handler(filename: str):
    handler = file_handlers.get(filename, None)
    if handler is None:
        handler = logging.FileHandler(filename)
        file_handlers[filename] = handler  # Am i need a lock?
    return handler


def get_file_logger(filename: str):
    """
    return a logger that writes records into a file.
    """
    logger = logging.getLogger(filename)
    handler = _get_file_logger_handler(filename)  # get singleton handler.
    handler.setFormatter(log_formatter)
    logger.addHandler(handler)  # each handler will be added only once.
    return logger


def get_bars(csv_file: str,
             symbol: str,
             exchange: Exchange,
             start_date: datetime = None,
             end_date: datetime = None, ):
    """
    获取bar
    数据存储目录: 项目/bar_data
    :param csv_file: csv文件路径
    :param symbol: 合约
    :param exchange 交易所
    :param start_date: datetime
    :param end_date: datetime
    :return:
    """
    bars = []

    import csv
    with open(file=csv_file, mode='r', encoding='utf8', newline='\n') as f:
        reader = csv.DictReader(f)

        count = 0

        for item in reader:

            dt = datetime.strptime(item['datetime'], '%Y-%m-%d %H:%M:%S')
            if start_date:
                if dt < start_date:
                    continue
            if end_date:
                if dt > end_date:
                    break

            bar = BarData(
                symbol=symbol,
                exchange=exchange,
                datetime=dt,
                interval=Interval.MINUTE,
                volume=float(item['volume']),
                open_price=float(item['open']),
                high_price=float(item['high']),
                low_price=float(item['low']),
                close_price=float(item['close']),
                open_interest=float(item['open_interest']),
                trading_day=item['trading_day'],
                gateway_name="Tdx",
            )

            bars.append(bar)

            # do some statistics
            count += 1

    return bars
