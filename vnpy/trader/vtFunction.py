# encoding: UTF-8

"""
包含一些开发中常用的函数
"""

import os,sys
import decimal
import json
from datetime import datetime,timedelta
import importlib
import re
from functools import lru_cache

MAX_NUMBER = 10000000000000
MAX_DECIMAL = 8


def import_module_by_str(import_module_name):
    """
    动态导入模块/函数
    :param import_module_name:
    :return:
    """
    import traceback
    from importlib import import_module, reload

    # 参数检查
    if len(import_module_name) == 0:
        print('import_module_by_str parameter error,return None')
        return None

    print('trying to import {}'.format(import_module_name))
    try:
        import_name = str(import_module_name).replace(':', '.')
        modules = import_name.split('.')
        if len(modules) == 1:
            mod = import_module(modules[0])
            return mod
        else:
            loaded_modules = '.'.join(modules[0:-1])
            print('import {}'.format(loaded_modules))
            mod = import_module(loaded_modules)

            comp = modules[-1]
            if not hasattr(mod, comp):
                loaded_modules = '.'.join([loaded_modules,comp])
                print('realod {}'.format(loaded_modules))
                mod = reload(loaded_modules)
            else:
                print('from {} import {}'.format(loaded_modules,comp))
                mod = getattr(mod, comp)
            return mod

    except Exception as ex:
        print('import {} fail,{},{}'.format(import_module_name,str(ex),traceback.format_exc()))

        return None



def floatToStr(float_str):
    """格式化显示浮点字符串，去除后面的0"""
    if '.' in float_str:
        llen = len(float_str)
        if llen > 0:
            for i in range(llen):
                vv = llen - i - 1
                if float_str[vv] not in ['.', '0']:
                    return float_str[:vv + 1]
                elif float_str[vv] in ['.']:
                    return float_str[:vv]
        return float_str
    else:
        return float_str

# ----------------------------------------------------------------------
def roundToPriceTick(priceTick, price):
    """取整价格到合约最小价格变动"""
    if not priceTick:
        return price

    if price > 0:
        # 根据最小跳动取整
        newPrice = price - price % priceTick
    else:
        # 兼容套利品种的负数价格
        newPrice = round(price / priceTick, 0) * priceTick

    # 数字货币，对浮点的长度有要求，需要砍除多余
    if isinstance(priceTick,float):
        price_exponent = decimal.Decimal(str(newPrice))
        tick_exponent = decimal.Decimal(str(priceTick))
        if abs(price_exponent.as_tuple().exponent) > abs(tick_exponent.as_tuple().exponent):
            newPrice = round(newPrice, ndigits=abs(tick_exponent.as_tuple().exponent))
            newPrice = float(str(newPrice))

    return newPrice

def roundToVolumeTick(volumeTick,volume):
    if volumeTick == 0:
        return volume
    # 取整
    newVolume = round(volume / volumeTick, 0) * volumeTick

    if isinstance(volumeTick,float):
        v_exponent = decimal.Decimal(str(volume))
        vt_exponent = decimal.Decimal(str(volumeTick))
        if abs(v_exponent.as_tuple().exponent) > abs(vt_exponent.as_tuple().exponent):
            newVolume = round(volume, ndigits=abs(vt_exponent.as_tuple().exponent))
            newVolume = float(str(newVolume))

    return newVolume

@lru_cache()
def getShortSymbol(symbol):
    """取得合约的短号"""
    # 套利合约
    if symbol.find(' ') != -1:
        # 排除SP SPC SPD
        s = symbol.split(' ')
        if len(s) < 2:
            return symbol
        symbol = s[1]

        # 只提取leg1合约
        if symbol.find('&') != -1:
            s = symbol.split('&')
            if len(s) < 2:
                return symbol
            symbol = s[0]

    p = re.compile(r"([A-Z]+)[0-9]+", re.I)
    shortSymbol = p.match(symbol)

    if shortSymbol is None:
        return symbol

    return shortSymbol.group(1)

@lru_cache()
def getFullSymbol(symbol):
    """
    获取全路径得合约名称
    """
    if symbol.endswith('SPD'):
        return symbol

    short_symbol = getShortSymbol(symbol)
    if short_symbol == symbol:
        return symbol

    symbol_month = symbol.replace(short_symbol, '')
    if len(symbol_month) == 3:
        if symbol_month[0] == '0':
            # 支持2020年合约
            return '{0}2{1}'.format(short_symbol, symbol_month)
        else:
            return '{0}1{1}'.format(short_symbol, symbol_month)
    else:
        return symbol

# -----------------------------------------
def systemSymbolToVnSymbol(symbol):
    """
    数字交易所合约symbol转化vnpy合约symbol
	ethusdt --> eth_usdt
	etcbtc  --> etc_btc
	ethusdt.HUOBI --> eth_usdt
	etcbtc.HUOBI  --> etc_btc
	ETHUSDT--> eth_usdt
    :param symbol:
    :return:
    """
    symbol = symbol.replace('_', '')
    symbol = ((symbol.split('.'))[0]).lower()
    if 'usdt' in symbol:
        return symbol[:-4] + "_usdt"
    else:
        return symbol[:-3] + "_" + symbol[-3:]

# -------------------------------------
def VnSymbolToSystemSymbol(symbol):
    """
    vnpy合约转化至数字交易所合约
    etc_btc  --> etcbtc
	eth_usdt --> ethusdt
	btc_usdt.OKEX => btcusdt
    :param symbol:
    :return:
    """
    symbol = (symbol.split('.'))[0]
    return (''.join(symbol.split('_'))).lower()


# ----------------------------------------------------------------------
def safeUnicode(value):
    """检查接口数据潜在的错误，保证转化为的字符串正确"""
    # 检查是数字接近0时会出现的浮点数上限
    if type(value) is int or type(value) is float:
        if value > MAX_NUMBER:
            value = 0
    
    # 检查防止小数点位过多
    if type(value) is float:
        d = decimal.Decimal(str(value))
        if abs(d.as_tuple().exponent) > MAX_DECIMAL:
            value = round(float(value), ndigits=MAX_DECIMAL)

    return value


def get_tdx_market_code(code):
    # 获取通达信股票的market code
    code = str(code)
    if code[0] in ['5', '6', '9'] or code[:3] in ["009", "126", "110", "201", "202", "203", "204"]:
        return 1
    return 0

#----------------------------------------------------------------------
def loadMongoSetting():
    """载入MongoDB数据库的配置"""

    try:
        from vnpy.trader.vtGlobal import globalSetting
        host = globalSetting['mongoHost']
        port = globalSetting['mongoPort']
        logging = globalSetting['mongoLogging']
    except:
        host = 'localhost'
        port = 27017
        logging = False
        
    return host, port, logging

#----------------------------------------------------------------------
def todayDate():
    """获取当前本机电脑时间的日期"""
    return datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)


def getTradingDate(dt=None):
    """
    根据输入的时间，返回交易日的日期
    :param dt:
    :return:
    """
    tradingDay = ''
    if dt is None:
        dt = datetime.now()

    if dt.hour >= 21:
        if dt.isoweekday() == 5:
            # 星期五=》星期一
            return (dt + timedelta(days=3)).strftime('%Y-%m-%d')
        else:
            # 第二天
            return (dt + timedelta(days=1)).strftime('%Y-%m-%d')
    elif dt.hour < 8 and dt.isoweekday() == 6:
        # 星期六=>星期一
        return (dt + timedelta(days=2)).strftime('%Y-%m-%d')
    else:
        return dt.strftime('%Y-%m-%d')


# 图标路径
iconPathDict = {}

path = os.path.abspath(os.path.join(os.path.dirname(__file__),'ico'))

for root, subdirs, files in os.walk(path):
    for fileName in files:
        if '.ico' in fileName:
            iconPathDict[fileName] = os.path.join(root, fileName)

# ----------------------------------------------------------------------
def loadIconPath(iconName):
    """加载程序图标路径"""
    global iconPathDict
    return iconPathDict.get(iconName, '')

# ----------------------------------------------------------------------
def getTempPath(name):
    """获取存放临时文件的路径"""
    tempPath = os.path.join(os.getcwd(), 'temp')
    if not os.path.exists(tempPath):
        os.makedirs(tempPath)

    path = os.path.join(tempPath, name)
    return path

def get_data_path():
    """获取存放数据文件的路径"""
    data_path = os.path.join(os.getcwd(),'data')
    if not os.path.exists(data_path):
        os.makedirs(data_path)
    return  data_path

# JSON配置文件路径
jsonPathDict = {}

# ----------------------------------------------------------------------
def getJsonPath(name, moduleFile):
    """
    获取JSON配置文件的路径：
    1. 优先从当前工作目录查找JSON文件
    2. 若无法找到则前往模块所在目录查找
    """
    currentFolder = os.getcwd()
    currentJsonPath = os.path.join(currentFolder, name)
    if os.path.isfile(currentJsonPath):
        print(u'use work folder file:{}'.format(currentJsonPath))
        jsonPathDict[name] = currentJsonPath
        return currentJsonPath

    moduleFolder = os.path.abspath(os.path.dirname(moduleFile))
    moduleJsonPath = os.path.join(moduleFolder, '.', name)
    print(u'use module file:{}'.format(moduleJsonPath))
    jsonPathDict[name] = moduleJsonPath
    return moduleJsonPath


def save_df_to_excel(file_name, sheet_name, df):
    """
    保存dataframe到execl
    :param file_name: 保存的excel文件名
    :param sheet_name: 保存的sheet
    :param df: dataframe
    :return: True/False
    """
    if file_name is None or sheet_name is None or df is None:
        return False

    # ----------------------------- 扩展的功能 ---------
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.drawing.image import Image
    except:
        print(u'can not import openpyxl', file=sys.stderr)

    if 'openpyxl' not in sys.modules:
        print(u'can not import openpyxl', file=sys.stderr)
        return False

    try:
        ws = None

        try:
            # 读取文件
            wb = openpyxl.load_workbook(file_name)
        except:
            # 创建一个excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
        try:
            # 定位WorkSheet
            if ws is None:
                ws = wb[sheet_name]
        except:
            # 创建一个WorkSheet
            ws = wb.create_sheet()
            ws.title = sheet_name

        rows = dataframe_to_rows(df)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Save the workbook
        wb.save(file_name)
        wb.close()
    except Exception as ex:
        import traceback
        print(u'save_df_to_excel exception:{}'.format(str(ex)), traceback.format_exc(),file=sys.stderr)

def save_text_to_excel(file_name, sheet_name, text):
    """
    保存文本文件到excel
    :param file_name:
    :param sheet_name:
    :param text:
    :return:
    """
    if file_name is None or len(sheet_name)==0 or len(text) == 0 :
        return False

    # ----------------------------- 扩展的功能 ---------
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.drawing.image import Image
    except:
        print(u'can not import openpyxl', file=sys.stderr)

    if 'openpyxl' not in sys.modules:
        return False

    try:
        ws = None
        try:
            # 读取文件
            wb = openpyxl.load_workbook(file_name)
        except:
            # 创建一个excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
        try:
            # 定位WorkSheet
            if ws is None:
                ws = wb[sheet_name]
        except:
            # 创建一个WorkSheet
            ws = wb.create_sheet()
            ws.title = sheet_name

        # 设置宽度，自动换行方式
        ws.column_dimensions["A"].width = 120
        ws['A2'].alignment = openpyxl.styles.Alignment(wrapText=True)
        ws['A2'].value = text

        # Save the workbook
        wb.save(file_name)
        wb.close()
        return True
    except Exception as ex:
        import traceback
        print(u'save_text_to_excel exception:{}'.format(str(ex)), traceback.format_exc(),file=sys.stderr)
        return False

def save_images_to_excel(file_name, sheet_name, image_names):
    """
    # 保存图形文件到excel
    :param file_name: excel文件名
    :param sheet_name: workSheet
    :param image_names: 图像文件名列表
    :return:
    """
    if file_name is None or len(sheet_name) == 0 or len(image_names) == 0:
        return False
    # ----------------------------- 扩展的功能 ---------
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.drawing.image import Image
    except:
        print(u'can not import openpyxl', file=sys.stderr)

    if 'openpyxl' not in sys.modules:
        return False
    try:
        ws = None

        try:
            # 读取文件
            wb = openpyxl.load_workbook(file_name)
        except:
            # 创建一个excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
        try:
            # 定位WorkSheet
            if ws is None:
                ws = wb[sheet_name]
        except:
            # 创建一个WorkSheet
            ws = wb.create_sheet()
            ws.title = sheet_name

        i = 1

        for image_name in image_names:
            try:
                # 加载图形文件
                img1 = Image(image_name)

                cell_id = 'A{0}'.format(i)
                ws[cell_id].value = image_name
                cell_id = 'A{0}'.format(i + 1)

                i += 30

                # 添加至对应的WorkSheet中
                ws.add_image(img1, cell_id)
            except:
                print('exception loading image {0}'.format(image_name), file=sys.stderr)
                return False

        # Save the workbook
        wb.save(file_name)
        wb.close()
        return True
    except Exception as ex:
        import traceback
        print(u'save_images_to_excel exception:{}'.format(str(ex)), traceback.format_exc(),file=sys.stderr)
        return False


def display_dual_axis(df, columns1, columns2=[], invert_yaxis1=False, invert_yaxis2=False, file_name=None, sheet_name=None,
             image_name=None):
    """
    显示(保存)双Y轴的走势图
    :param df: DataFrame
    :param columns1:  y1轴
    :param columns2: Y2轴
    :param invert_yaxis1: Y1 轴反转
    :param invert_yaxis2: Y2 轴翻转
    :param file_name:   保存的excel 文件名称
    :param sheet_name:  excel 的sheet
    :param image_name:  保存的image 文件名
    :return:
    """

    import matplotlib
    import matplotlib.pyplot as plt
    matplotlib.rcParams['figure.figsize'] = (20.0, 10.0)

    df1 = df[columns1]
    df1.index = list(range(len(df)))
    fig, ax1 = plt.subplots()
    if invert_yaxis1:
        ax1.invert_yaxis()
    ax1.plot(df1)

    if len(columns2) > 0:
        df2 = df[columns2]
        df2.index = list(range(len(df)))
        ax2 = ax1.twinx()
        if invert_yaxis2:
            ax2.invert_yaxis()
        ax2.plot(df2)

    # 修改x轴得label为时间
    xt = ax1.get_xticks()
    xt2 = [df.index[int(i)] for i in xt[1:-2]]
    xt2.insert(0, '')
    xt2.append('')
    ax1.set_xticklabels(xt2)

    # 是否保存图片到文件
    if image_name is not None:
        fig = plt.gcf()
        fig.savefig(image_name, bbox_inches='tight')

        # 插入图片到指定的excel文件sheet中并保存excel
        if file_name is not None and sheet_name is not None:
            save_images_to_excel(file_name, sheet_name, [image_name])
    else:
        plt.show()